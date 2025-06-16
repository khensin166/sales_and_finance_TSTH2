# finance/views.py
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import get_template
from django.utils.dateparse import parse_date
from xhtml2pdf import pisa # pylint: disable=import-error
import openpyxl # pylint: disable=import-error
from openpyxl.styles import Font, Alignment # pylint: disable=import-error
from openpyxl.utils import get_column_letter # pylint: disable=import-error

from .models import Finance
from .filters import FinanceFilter

def export_finance_pdf(request):
    filterset = FinanceFilter(request.GET, queryset=Finance.objects.all())
    if not filterset.is_valid():
        return HttpResponse(f"Filter tidak valid: {filterset.errors}", status=400)

    finances = filterset.qs

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    formatted_start = formatted_end = None
    if start_date_str and end_date_str:
        try:
            start_date = datetime.combine(parse_date(start_date_str), datetime.min.time())
            end_date = datetime.combine(parse_date(end_date_str), datetime.max.time())
            formatted_start = start_date.strftime("%d %b %Y %H:%M")
            formatted_end = end_date.strftime("%d %b %Y %H:%M")
        except:
            pass

    template = get_template("pdf/finance_report.html")
    html = template.render({
        "finances": finances,
        "start_date": formatted_start,
        "end_date": formatted_end
    })

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=finance_report.pdf"
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Gagal membuat PDF", status=500)
    return response


def export_finance_excel(request):
    filterset = FinanceFilter(request.GET, queryset=Finance.objects.all())
    if not filterset.is_valid():
        return HttpResponse(f"Filter tidak valid: {filterset.errors}", status=400)

    finances = filterset.qs

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    formatted_start = formatted_end = None
    if start_date_str and end_date_str:
        try:
            start_date = datetime.combine(parse_date(start_date_str), datetime.min.time())
            end_date = datetime.combine(parse_date(end_date_str), datetime.max.time())
            formatted_start = start_date.strftime("%d %b %Y %H:%M")
            formatted_end = end_date.strftime("%d %b %Y %H:%M")
        except:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finance Report"

    ws.merge_cells('A1:E1')
    ws['A1'] = f"Export period: {formatted_start} - {formatted_end}" if formatted_start and formatted_end else "All Data"
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = ['#', 'Date & Time', 'Transaction Type', 'Description', 'Amount']
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for idx, finance in enumerate(finances, start=1):
        ws.append([
            idx,
            finance.transaction_date.strftime("%d %b %Y %H:%M"),
            finance.transaction_type,
            finance.description,
            float(finance.amount)
        ])

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=finance_report.xlsx'
    wb.save(response)
    return response
