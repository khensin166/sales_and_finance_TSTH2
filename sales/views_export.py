# Standard Library
from datetime import datetime

# Third-party Libraries
from xhtml2pdf import pisa # pylint: disable=import-error
import openpyxl # pylint: disable=import-error
from openpyxl.styles import Font, Alignment # pylint: disable=import-error
from openpyxl.utils import get_column_letter # pylint: disable=import-error

# Django Core
from django.http import HttpResponse
from django.template.loader import get_template
from django.utils.dateparse import parse_date

# Local App
from .models import SalesTransaction
from .filters import SalesTransactionFilter

import io
import logging
from openpyxl import Workbook

logger = logging.getLogger(__name__)

def export_pdf(request):
    filterset = SalesTransactionFilter(request.GET, queryset=SalesTransaction.objects.all())
    if not filterset.is_valid():
        return HttpResponse(f"Filter tidak valid: {filterset.errors}", status=400)

    transactions = filterset.qs

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    formatted_start = formatted_end = None
    if start_date_str and end_date_str:
        try:
            start_date = datetime.combine(parse_date(start_date_str), datetime.min.time())
            end_date = datetime.combine(parse_date(end_date_str), datetime.max.time())
            formatted_start = start_date.strftime("%d %b %Y %H:%M")
            formatted_end = end_date.strftime("%d %b %Y %H:%M")
        except:
            pass

    total_quantity = sum(t.quantity for t in transactions)
    total_sales = sum(t.total_price for t in transactions)
    context = {
        'transactions': transactions,
        'total_quantity': total_quantity,
        'total_sales': total_sales,
        'start_date': formatted_start,
        'end_date': formatted_end,
    }

    template = get_template("pdf/sales_transaction_report.html")
    html = template.render(context)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=sales_transaction.pdf"
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Gagal membuat PDF", status=500)
    return response

def export_excel(request):
    try:
        # Filter sales transactions
        filterset = SalesTransactionFilter(request.GET, queryset=SalesTransaction.objects.all())
        if not filterset.is_valid():
            logger.error("Invalid filter: %s", filterset.errors)
            return HttpResponse(f"Invalid filter: {filterset.errors}", status=400)

        transactions = filterset.qs

        # Parse date range
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        formatted_start = formatted_end = None
        if start_date_str and end_date_str:
            try:
                start_date = datetime.combine(parse_date(start_date_str), datetime.min.time())
                end_date = datetime.combine(parse_date(end_date_str), datetime.max.time())
                formatted_start = start_date.strftime("%d %b %Y %H:%M")
                formatted_end = end_date.strftime("%d %b %Y %H:%M")
            except ValueError as e:
                logger.warning(f"Invalid date format: {start_date_str}, {end_date_str}. Error: {str(e)}")
                # Continue with no date range in header

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Transactions"

        # Add header
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Export period: {formatted_start} - {formatted_end}" if formatted_start and formatted_end else "All Data"
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Add column headers
        headers = ['#', 'Transaction Date', 'Order No', 'Customer Name', 'Quantity', 'Total Price', 'Payment Method']
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add transaction data
        for idx, transaction in enumerate(transactions, start=1):
            try:
                ws.append([
                    idx,
                    transaction.transaction_date.strftime("%d %b %Y %H:%M"),
                    transaction.order.order_no,
                    transaction.order.customer_name,
                    transaction.quantity,
                    float(transaction.total_price),
                    transaction.payment_method or "-"
                ])
            except AttributeError as e:
                logger.error(f"Error processing transaction {idx}: {str(e)}")
                continue

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except TypeError:
                    continue
            ws.column_dimensions[column].width = max_length + 2

        # Save workbook to a buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create response
        response = HttpResponse(
            content=buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=sales_transactions_{start_date_str}_to_{end_date_str}.xlsx'
        buffer.close()

        logger.info("Excel export successful for %s to %s", start_date_str, end_date_str)
        return response

    except Exception as e:
        logger.error("Unexpected error in export_excel: %s", str(e))
        return HttpResponse(f"Error generating Excel file: {str(e)}", status=500)