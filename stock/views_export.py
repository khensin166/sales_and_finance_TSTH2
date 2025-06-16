# Standard Library
from datetime import datetime

# Third-party Libraries
from xhtml2pdf import pisa  # pylint: disable=import-error
import openpyxl  # pylint: disable=import-error
from openpyxl.styles import Font, Alignment  # pylint: disable=import-error
from openpyxl.utils import get_column_letter  # pylint: disable=import-error

# Django Core
from django.http import HttpResponse
from django.template.loader import get_template
from django.utils.dateparse import parse_date

# Local App
from .models import StockHistory
from .filters import StockHistoryFilter


def export_pdf(request):
    filterset = StockHistoryFilter(request.GET, queryset=StockHistory.objects.all())
    if not filterset.is_valid():
        return HttpResponse(f"Filter tidak valid: {filterset.errors}", status=400)

    histories = filterset.qs

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    formatted_start = formatted_end = None
    if start_date_str and end_date_str:
        try:
            start_date = datetime.combine(parse_date(start_date_str), datetime.min.time())
            end_date = datetime.combine(parse_date(end_date_str), datetime.max.time())
            formatted_start = start_date.strftime("%d %b %Y %H:%M")
            formatted_end = end_date.strftime("%d %b %Y %H:%M")
        except:
            pass

    template = get_template("pdf/history_report.html")
    html = template.render({
        "histories": histories,
        "start_date": formatted_start,
        "end_date": formatted_end
    })

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=product_history.pdf"
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("Gagal membuat PDF", status=500)
    return response


def export_excel(request):
    filterset = StockHistoryFilter(request.GET, queryset=StockHistory.objects.all())
    if not filterset.is_valid():
        return HttpResponse(f"Filter tidak valid: {filterset.errors}", status=400)

    histories = filterset.qs

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    formatted_start = formatted_end = None
    if start_date_str and end_date_str:
        try:
            start_date = datetime.combine(parse_date(start_date_str), datetime.min.time())
            end_date = datetime.combine(parse_date(end_date_str), datetime.max.time())
            formatted_start = start_date.strftime("%d %b %Y %H:%M")
            formatted_end = end_date.strftime("%d %b %Y %H:%M")
        except:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product History"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Export period: {formatted_start} - {formatted_end}" if formatted_start and formatted_end else "All Data"
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    headers = ['#', 'Date & Time', 'Product Type', 'Change Type', 'Quantity', 'Total Price']
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for idx, history in enumerate(histories, start=1):
        product_type = history.product_stock.product_type.product_name if history.product_stock and history.product_stock.product_type else "-"
        unit = history.product_stock.product_type.unit if history.product_stock and history.product_stock.product_type else ""
        quantity = f"{history.quantity_change} {unit}" if unit else history.quantity_change

        ws.append([
            idx,
            history.change_date.strftime("%d %b %Y %H:%M"),
            product_type,
            history.change_type,
            quantity,
            float(history.total_price)
        ])

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=product_history.xlsx'
    wb.save(response)
    return response
